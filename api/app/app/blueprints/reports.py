from flask import Blueprint, render_template, request, jsonify, send_file
from flask_login import login_required
from datetime import datetime, timedelta
from sqlalchemy import func, and_, desc
from ..models import Ticket, TimeEntry, User, Client, Service, ServiceOrder
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import io
import os

bp = Blueprint("reports", __name__)


@bp.route("/")
@login_required
def index():
	"""Página principal de relatórios com abas"""
	users = User.query.filter_by(status="1").order_by(User.name.asc()).all()
	clients = Client.query.all()
	services = Service.query.all()
	
	# Dados para filtros padrão
	today = datetime.now().date()
	start_of_month = today.replace(day=1)
	
	return render_template("reports/index.html", 
		users=users, 
		clients=clients, 
		services=services,
		default_start=start_of_month.isoformat(),
		default_end=today.isoformat()
	)


@bp.route("/api/hours-by-client")
@login_required
def api_hours_by_client():
	"""API para relatório de horas por cliente"""
	start = request.args.get("start")
	end = request.args.get("end")
	
	query = TimeEntry.query.join(Ticket, TimeEntry.ticket_id == Ticket.id)
	
	# Filtrar apenas tickets válidos (não cancelados)
	query = query.filter(Ticket.status != "cancelado")
	
	if start:
		start_dt = datetime.fromisoformat(start)
		query = query.filter(TimeEntry.created_at >= start_dt)
	if end:
		# Corrigir: adicionar 23:59:59 para incluir todo o dia
		end_dt = datetime.fromisoformat(end) + timedelta(hours=23, minutes=59, seconds=59)
		query = query.filter(TimeEntry.created_at <= end_dt)
	
	# Agrupar por cliente
	results = (
		query.with_entities(
			Ticket.client_id,
			Ticket.external_client_id,
			Ticket.external_client_name,
			func.sum(TimeEntry.hours).label("total_hours"),
			func.count(Ticket.id.distinct()).label("tickets_count"),
			func.avg(TimeEntry.hours).label("avg_hours_per_entry")
		)
		.group_by(Ticket.client_id, Ticket.external_client_id, Ticket.external_client_name)
		.order_by(desc("total_hours"))
		.all()
	)
	
	# Buscar clientes internos
	internal_clients = {c.id: c for c in Client.query.all()}
	
	data = []
	for client_id, ext_client_id, ext_client_name, total_hours, tickets_count, avg_hours in results:
		if client_id and client_id in internal_clients:
			client_name = internal_clients[client_id].name
			client_type = "Interno"
		elif ext_client_name:
			client_name = ext_client_name
			client_type = "Externo"
		else:
			client_name = "Cliente não encontrado"
			client_type = "Desconhecido"
		
		data.append({
			"client_id": client_id,
			"external_client_id": ext_client_id,
			"external_client_name": ext_client_name,
			"client_name": client_name,
			"client_type": client_type,
			"total_hours": float(total_hours or 0),
			"tickets_count": tickets_count or 0,
			"avg_hours_per_entry": float(avg_hours or 0)
		})
	
	return jsonify(data)


@bp.route("/api/hours-by-client-detailed")
@login_required
def api_hours_by_client_detailed():
	"""API para relatório detalhado de horas por cliente específico"""
	start = request.args.get("start")
	end = request.args.get("end")
	client_id = request.args.get("client_id", type=int)
	external_client_id = request.args.get("external_client_id", type=int)
	external_client_name = request.args.get("external_client_name")
	
	if not client_id and not external_client_id and not external_client_name:
		return jsonify({"error": "Cliente não especificado"}), 400
	
	# Buscar tickets do cliente
	tickets_query = Ticket.query
	
	if client_id:
		tickets_query = tickets_query.filter(Ticket.client_id == client_id)
	elif external_client_id:
		tickets_query = tickets_query.filter(Ticket.external_client_id == external_client_id)
	elif external_client_name:
		tickets_query = tickets_query.filter(Ticket.external_client_name == external_client_name)
	
	# Filtrar apenas tickets válidos (não cancelados)
	tickets_query = tickets_query.filter(Ticket.status != "cancelado")
	
	if start:
		start_dt = datetime.fromisoformat(start)
		tickets_query = tickets_query.filter(Ticket.created_at >= start_dt)
	if end:
		end_dt = datetime.fromisoformat(end) + timedelta(hours=23, minutes=59, seconds=59)
		tickets_query = tickets_query.filter(Ticket.created_at <= end_dt)
	
	tickets = tickets_query.order_by(Ticket.created_at.desc()).all()
	
	# Buscar todas as entradas de tempo dos tickets
	ticket_ids = [t.id for t in tickets]
	time_entries = []
	if ticket_ids:
		time_entries = TimeEntry.query.filter(TimeEntry.ticket_id.in_(ticket_ids)).order_by(TimeEntry.created_at.desc()).all()
	
	# Calcular totais
	total_hours = sum(entry.hours for entry in time_entries)
	total_tickets = len(tickets)
	total_entries = len(time_entries)
	
	# Determinar informações do cliente
	client_info = {}
	if client_id:
		client = Client.query.get(client_id)
		if client:
			client_info = {
				"name": client.name,
				"type": "Interno",
				"document": client.document,
				"phone": client.phone
			}
	elif external_client_name:
		client_info = {
			"name": external_client_name,
			"type": "Externo",
			"document": None,
			"phone": None
		}
	
	# Formatar dados dos tickets com suas entradas de tempo
	tickets_data = []
	for ticket in tickets:
		ticket_entries = [entry for entry in time_entries if entry.ticket_id == ticket.id]
		
		tickets_data.append({
			"id": ticket.id,
			"title": ticket.title,
			"description": ticket.description,
			"status": ticket.status,
			"created_at": ticket.created_at.isoformat() if ticket.created_at else None,
			"closed_at": ticket.closed_at.isoformat() if ticket.closed_at else None,
			"service_name": ticket.service.name if ticket.service else "N/A",
			"assigned_to": ticket.assigned_to_user.name if ticket.assigned_to_user else "Não atribuído",
			"total_hours": sum(entry.hours for entry in ticket_entries),
			"time_entries": [
				{
					"id": entry.id,
					"description": entry.comment or "Sem descrição",
					"hours": entry.hours,
					"created_at": entry.created_at.isoformat() if entry.created_at else None,
					"user_name": entry.user.name if entry.user else "N/A",
					"no_charge": entry.no_charge
				}
				for entry in ticket_entries
			]
		})
	
	return jsonify({
		"client_info": client_info,
		"period": {
			"start": start,
			"end": end
		},
		"totals": {
			"total_hours": total_hours,
			"total_tickets": total_tickets,
			"total_entries": total_entries
		},
		"tickets": tickets_data
	})


@bp.route("/api/hours-by-client-synthetic")
@login_required
def api_hours_by_client_synthetic():
	"""API para relatório sintético de horas por cliente específico"""
	start = request.args.get("start")
	end = request.args.get("end")
	client_id = request.args.get("client_id", type=int)
	external_client_id = request.args.get("external_client_id", type=int)
	external_client_name = request.args.get("external_client_name")
	
	if not client_id and not external_client_id and not external_client_name:
		return jsonify({"error": "Cliente não especificado"}), 400
	
	# Buscar tickets do cliente
	tickets_query = Ticket.query
	
	if client_id:
		tickets_query = tickets_query.filter(Ticket.client_id == client_id)
	elif external_client_id:
		tickets_query = tickets_query.filter(Ticket.external_client_id == external_client_id)
	elif external_client_name:
		tickets_query = tickets_query.filter(Ticket.external_client_name == external_client_name)
	
	# Filtrar apenas tickets válidos (não cancelados)
	tickets_query = tickets_query.filter(Ticket.status != "cancelado")
	
	if start:
		start_dt = datetime.fromisoformat(start)
		tickets_query = tickets_query.filter(Ticket.created_at >= start_dt)
	if end:
		end_dt = datetime.fromisoformat(end) + timedelta(hours=23, minutes=59, seconds=59)
		tickets_query = tickets_query.filter(Ticket.created_at <= end_dt)
	
	tickets = tickets_query.order_by(Ticket.created_at.desc()).all()
	
	# Calcular totais
	total_hours = sum(ticket.total_hours() for ticket in tickets)
	total_tickets = len(tickets)
	
	# Determinar informações do cliente
	client_info = {}
	if client_id:
		client = Client.query.get(client_id)
		if client:
			client_info = {
				"name": client.name,
				"type": "Interno",
				"document": client.document,
				"phone": client.phone
			}
	elif external_client_name:
		client_info = {
			"name": external_client_name,
			"type": "Externo",
			"document": None,
			"phone": None
		}
	
	# Formatar dados sintéticos dos tickets
	tickets_data = []
	for ticket in tickets:
		tickets_data.append({
			"id": ticket.id,
			"title": ticket.title,
			"status": ticket.status,
			"created_at": ticket.created_at.isoformat() if ticket.created_at else None,
			"closed_at": ticket.closed_at.isoformat() if ticket.closed_at else None,
			"service_name": ticket.service.name if ticket.service else "N/A",
			"assigned_to": ticket.assigned_to_user.name if ticket.assigned_to_user else "Não atribuído",
			"total_hours": ticket.total_hours()
		})
	
	return jsonify({
		"client_info": client_info,
		"period": {
			"start": start,
			"end": end
		},
		"totals": {
			"total_hours": total_hours,
			"total_tickets": total_tickets
		},
		"tickets": tickets_data
	})


@bp.route("/api/hours-by-technician")
@login_required
def api_hours_by_technician():
	"""API para relatório de horas por técnico"""
	start = request.args.get("start")
	end = request.args.get("end")
	
	query = TimeEntry.query.join(User, TimeEntry.user_id == User.id)
	
	# Filtrar apenas tickets válidos (não cancelados)
	query = query.join(Ticket, TimeEntry.ticket_id == Ticket.id).filter(Ticket.status != "cancelado")
	query = query.filter(User.status == "1")
	
	if start:
		start_dt = datetime.fromisoformat(start)
		query = query.filter(TimeEntry.created_at >= start_dt)
	if end:
		# Corrigir: adicionar 23:59:59 para incluir todo o dia
		end_dt = datetime.fromisoformat(end) + timedelta(hours=23, minutes=59, seconds=59)
		query = query.filter(TimeEntry.created_at <= end_dt)
	
	results = (
		query.with_entities(
			TimeEntry.user_id,
			User.name,
			User.role,
			func.sum(TimeEntry.hours).label("total_hours"),
			func.count(TimeEntry.id).label("entries_count"),
			func.count(Ticket.id.distinct()).label("tickets_count"),
			func.avg(TimeEntry.hours).label("avg_hours_per_entry")
		)
		.group_by(TimeEntry.user_id, User.name, User.role)
		.order_by(desc("total_hours"))
		.all()
	)
	
	data = []
	for user_id, name, role, total_hours, entries_count, tickets_count, avg_hours in results:
		data.append({
			"user_id": user_id,
			"name": name,
			"role": role,
			"total_hours": float(total_hours or 0),
			"entries_count": entries_count or 0,
			"tickets_count": tickets_count or 0,
			"avg_hours_per_entry": float(avg_hours or 0)
		})
	
	return jsonify(data)


@bp.route("/api/billing-by-technician")
@login_required
def api_billing_by_technician():
	"""API para relatório de faturamento por técnico (tickets + ordens de serviço)"""
	start = request.args.get("start")
	end = request.args.get("end")
	
	# 1. FATURAMENTO DE TICKETS
	tickets_query = Ticket.query.join(User, Ticket.assigned_to_id == User.id)
	
	# Filtrar apenas tickets válidos (não cancelados) e fechados (que têm faturamento)
	tickets_query = tickets_query.filter(Ticket.status != "cancelado")
	tickets_query = tickets_query.filter(Ticket.status == "fechado")  # Só tickets fechados têm faturamento
	tickets_query = tickets_query.filter(Ticket.closed_at.isnot(None))  # Garantir que tem data de fechamento
	tickets_query = tickets_query.filter(User.status == "1")
	
	if start:
		start_dt = datetime.fromisoformat(start)
		tickets_query = tickets_query.filter(Ticket.closed_at >= start_dt)
	if end:
		end_dt = datetime.fromisoformat(end) + timedelta(hours=23, minutes=59, seconds=59)
		tickets_query = tickets_query.filter(Ticket.closed_at <= end_dt)
	
	# Agregação de tickets por técnico
	tickets_results = (
		tickets_query.with_entities(
			Ticket.assigned_to_id,
			User.name,
			User.role,
			func.sum(Ticket.total_cost).label("tickets_billing"),
			func.count(Ticket.id).label("tickets_count")
		)
		.group_by(Ticket.assigned_to_id, User.name, User.role)
		.all()
	)
	
	# 2. FATURAMENTO DE ORDENS DE SERVIÇO
	service_orders_query = ServiceOrder.query.join(User, ServiceOrder.technician_id == User.id)
	
	# Filtrar apenas ordens com valor > 0 (que geram faturamento)
	service_orders_query = service_orders_query.filter(ServiceOrder.value > 0)
	service_orders_query = service_orders_query.filter(User.status == "1")
	
	if start:
		start_dt = datetime.fromisoformat(start)
		service_orders_query = service_orders_query.filter(ServiceOrder.completion_date >= start_dt)
	if end:
		end_dt = datetime.fromisoformat(end) + timedelta(hours=23, minutes=59, seconds=59)
		service_orders_query = service_orders_query.filter(ServiceOrder.completion_date <= end_dt)
	
	# Agregação de ordens de serviço por técnico
	service_orders_results = (
		service_orders_query.with_entities(
			ServiceOrder.technician_id,
			User.name,
			User.role,
			func.sum(ServiceOrder.value).label("service_orders_billing"),
			func.count(ServiceOrder.id).label("service_orders_count")
		)
		.group_by(ServiceOrder.technician_id, User.name, User.role)
		.all()
	)
	
	# 3. COMBINAR RESULTADOS
	# Criar dicionários para facilitar a combinação
	tickets_dict = {}
	for user_id, name, role, tickets_billing, tickets_count in tickets_results:
		tickets_dict[user_id] = {
			"name": name,
			"role": role,
			"tickets_billing": float(tickets_billing or 0),
			"tickets_count": tickets_count or 0
		}
	
	service_orders_dict = {}
	for user_id, name, role, service_orders_billing, service_orders_count in service_orders_results:
		service_orders_dict[user_id] = {
			"name": name,
			"role": role,
			"service_orders_billing": float(service_orders_billing or 0),
			"service_orders_count": service_orders_count or 0
		}
	
	# Combinar dados de todos os usuários únicos
	all_user_ids = set(tickets_dict.keys()) | set(service_orders_dict.keys())
	
	data = []
	for user_id in all_user_ids:
		tickets_data = tickets_dict.get(user_id, {"tickets_billing": 0, "tickets_count": 0})
		service_orders_data = service_orders_dict.get(user_id, {"service_orders_billing": 0, "service_orders_count": 0})
		
		# Usar dados do usuário de qualquer uma das fontes (devem ser iguais)
		user_data = tickets_data if user_id in tickets_dict else service_orders_data
		
		total_billing = tickets_data["tickets_billing"] + service_orders_data["service_orders_billing"]
		
		data.append({
			"user_id": user_id,
			"name": user_data["name"],
			"role": user_data["role"],
			"total_billing": total_billing,
			"tickets_billing": tickets_data["tickets_billing"],
			"service_orders_billing": service_orders_data["service_orders_billing"],
			"tickets_count": tickets_data["tickets_count"],
			"service_orders_count": service_orders_data["service_orders_count"],
			"total_items": tickets_data["tickets_count"] + service_orders_data["service_orders_count"]
		})
	
	# Ordenar por faturamento total decrescente
	data.sort(key=lambda x: x["total_billing"], reverse=True)
	
	return jsonify(data)


@bp.route("/api/billing-tickets-by-technician/<int:technician_id>")
@login_required
def api_billing_tickets_by_technician(technician_id):
	"""API para buscar tickets faturados de um técnico específico"""
	start = request.args.get("start")
	end = request.args.get("end")
	
	# Buscar tickets fechados do técnico
	tickets_query = Ticket.query.join(User, Ticket.assigned_to_id == User.id)
	tickets_query = tickets_query.filter(Ticket.assigned_to_id == technician_id)
	tickets_query = tickets_query.filter(Ticket.status == "fechado")
	tickets_query = tickets_query.filter(Ticket.closed_at.isnot(None))
	tickets_query = tickets_query.filter(Ticket.total_cost > 0)  # Apenas tickets com faturamento
	
	if start:
		start_dt = datetime.fromisoformat(start)
		tickets_query = tickets_query.filter(Ticket.closed_at >= start_dt)
	if end:
		end_dt = datetime.fromisoformat(end) + timedelta(hours=23, minutes=59, seconds=59)
		tickets_query = tickets_query.filter(Ticket.closed_at <= end_dt)
	
	tickets = tickets_query.order_by(Ticket.closed_at.desc()).all()
	
	# Buscar ordens de serviço do técnico
	service_orders_query = ServiceOrder.query.join(User, ServiceOrder.technician_id == User.id)
	service_orders_query = service_orders_query.filter(ServiceOrder.technician_id == technician_id)
	service_orders_query = service_orders_query.filter(ServiceOrder.value > 0)
	
	if start:
		start_dt = datetime.fromisoformat(start)
		service_orders_query = service_orders_query.filter(ServiceOrder.completion_date >= start_dt)
	if end:
		end_dt = datetime.fromisoformat(end) + timedelta(hours=23, minutes=59, seconds=59)
		service_orders_query = service_orders_query.filter(ServiceOrder.completion_date <= end_dt)
	
	service_orders = service_orders_query.order_by(ServiceOrder.completion_date.desc()).all()
	
	# Formatar dados dos tickets
	tickets_data = []
	for ticket in tickets:
		tickets_data.append({
			"id": ticket.id,
			"title": ticket.title,
			"client_name": ticket.display_client_name(),
			"service_name": ticket.service.name if ticket.service else "N/A",
			"total_cost": float(ticket.total_cost),
			"total_hours": ticket.total_hours(),
			"closed_at": ticket.closed_at.isoformat() if ticket.closed_at else None,
			"type": "ticket"
		})
	
	# Formatar dados das ordens de serviço
	service_orders_data = []
	for order in service_orders:
		service_orders_data.append({
			"id": order.id,
			"codigo": order.codigo,
			"client_name": order.client_name,
			"service_executed": order.service_executed,
			"value": float(order.value),
			"completion_date": order.completion_date.isoformat() if order.completion_date else None,
			"type": "service_order"
		})
	
	return jsonify({
		"tickets": tickets_data,
		"service_orders": service_orders_data,
		"total_tickets": len(tickets_data),
		"total_service_orders": len(service_orders_data)
	})


@bp.route("/api/tickets-by-technician")
@login_required
def api_tickets_by_technician():
	"""API para relatório de tickets por técnico"""
	start = request.args.get("start")
	end = request.args.get("end")
	
	query = Ticket.query.join(User, Ticket.assigned_to_id == User.id).filter(User.status == "1")
	
	if start:
		start_dt = datetime.fromisoformat(start)
		query = query.filter(Ticket.created_at >= start_dt)
	if end:
		# Corrigir: adicionar 23:59:59 para incluir todo o dia
		end_dt = datetime.fromisoformat(end) + timedelta(hours=23, minutes=59, seconds=59)
		query = query.filter(Ticket.created_at <= end_dt)
	
	# Separar contagem de tickets e horas para evitar duplicação
	tickets_results = (
		query.with_entities(
			Ticket.assigned_to_id,
			User.name,
			User.role,
			func.count(Ticket.id).label("total_tickets"),
			func.count(Ticket.id).filter(Ticket.status == "aberto").label("open_tickets"),
			func.count(Ticket.id).filter(Ticket.status == "em_andamento").label("in_progress_tickets"),
			func.count(Ticket.id).filter(Ticket.status == "fechado").label("closed_tickets"),
			func.count(Ticket.id).filter(Ticket.status == "cancelado").label("cancelled_tickets")
		)
		.group_by(Ticket.assigned_to_id, User.name, User.role)
		.all()
	)
	
	# Calcular horas separadamente
	hours_results = (
		query.with_entities(
			Ticket.assigned_to_id,
			func.sum(TimeEntry.hours).label("total_hours")
		)
		.outerjoin(TimeEntry, Ticket.id == TimeEntry.ticket_id)
		.filter(Ticket.status != "cancelado")
		.group_by(Ticket.assigned_to_id)
		.all()
	)
	
	# Combinar resultados
	hours_dict = {user_id: hours for user_id, hours in hours_results}
	results = []
	for user_id, name, role, total_tickets, open_tickets, in_progress_tickets, closed_tickets, cancelled_tickets in tickets_results:
		total_hours = hours_dict.get(user_id, 0) or 0
		results.append((user_id, name, role, total_tickets, open_tickets, in_progress_tickets, closed_tickets, cancelled_tickets, total_hours))
	
	data = []
	for user_id, name, role, total_tickets, open_tickets, in_progress_tickets, closed_tickets, cancelled_tickets, total_hours in results:
		data.append({
			"user_id": user_id,
			"name": name,
			"role": role,
			"total_tickets": total_tickets or 0,
			"open_tickets": open_tickets or 0,
			"in_progress_tickets": in_progress_tickets or 0,
			"closed_tickets": closed_tickets or 0,
			"cancelled_tickets": cancelled_tickets or 0,
			"total_hours": float(total_hours or 0),
			"avg_resolution_time": 0  # Removido temporariamente
		})
	
	return jsonify(data)


@bp.route("/api/tickets-by-client")
@login_required
def api_tickets_by_client():
	"""API para relatório de tickets por cliente."""
	start = request.args.get("start")
	end = request.args.get("end")

	query = Ticket.query
	if start:
		query = query.filter(Ticket.created_at >= datetime.fromisoformat(start))
	if end:
		end_dt = datetime.fromisoformat(end) + timedelta(hours=23, minutes=59, seconds=59)
		query = query.filter(Ticket.created_at <= end_dt)

	tickets_results = (
		query.with_entities(
			Ticket.client_id,
			Ticket.external_client_id,
			Ticket.external_client_name,
			func.count(Ticket.id).label("total_tickets"),
			func.count(Ticket.id).filter(Ticket.status == "aberto").label("open_tickets"),
			func.count(Ticket.id).filter(Ticket.status == "em_andamento").label("in_progress_tickets"),
			func.count(Ticket.id).filter(Ticket.status == "fechado").label("closed_tickets"),
			func.count(Ticket.id).filter(Ticket.status == "cancelado").label("cancelled_tickets"),
		)
		.group_by(Ticket.client_id, Ticket.external_client_id, Ticket.external_client_name)
		.all()
	)

	hours_results = (
		query.with_entities(
			Ticket.client_id,
			Ticket.external_client_id,
			Ticket.external_client_name,
			func.sum(TimeEntry.hours).label("total_hours"),
		)
		.outerjoin(TimeEntry, Ticket.id == TimeEntry.ticket_id)
		.filter(Ticket.status != "cancelado")
		.group_by(Ticket.client_id, Ticket.external_client_id, Ticket.external_client_name)
		.all()
	)
	hours_dict = {
		(cid, eid, ename): hours
		for cid, eid, ename, hours in hours_results
	}

	internal_clients = {c.id: c for c in Client.query.all()}
	data = []
	for client_id, ext_id, ext_name, total, open_n, progress_n, closed_n, cancelled_n in tickets_results:
		if client_id and client_id in internal_clients:
			client_name = internal_clients[client_id].name
			client_type = "Interno"
		elif ext_name:
			client_name = ext_name
			client_type = "Externo"
		else:
			client_name = "Sem cliente"
			client_type = "—"
		data.append({
			"client_id": client_id,
			"external_client_id": ext_id,
			"external_client_name": ext_name,
			"client_name": client_name,
			"client_type": client_type,
			"total_tickets": total or 0,
			"open_tickets": open_n or 0,
			"in_progress_tickets": progress_n or 0,
			"closed_tickets": closed_n or 0,
			"cancelled_tickets": cancelled_n or 0,
			"total_hours": float(hours_dict.get((client_id, ext_id, ext_name), 0) or 0),
		})
	data.sort(key=lambda r: r["total_tickets"], reverse=True)
	return jsonify(data)


@bp.route("/api/productivity-metrics")
@login_required
def api_productivity_metrics():
	"""API para métricas de produtividade"""
	start = request.args.get("start")
	end = request.args.get("end")
	
	if start:
		start_dt = datetime.fromisoformat(start)
	else:
		start_dt = datetime.now() - timedelta(days=30)
	
	if end:
		end_dt = datetime.fromisoformat(end)
		# YYYY-MM-DD chega como meia-noite; inclui o dia inteiro (igual /api/web/reports).
		if len((end or "").strip()) <= 10:
			end_dt = end_dt.replace(hour=23, minute=59, second=59)
	else:
		end_dt = datetime.now()
	
	# Tickets por dia
	daily_tickets = (
		Ticket.query.with_entities(
			func.date(Ticket.created_at).label("date"),
			func.count(Ticket.id).label("tickets_count")
		)
		.filter(
			Ticket.created_at >= start_dt,
			Ticket.created_at <= end_dt,
			Ticket.status != "cancelado",
		)
		.group_by(func.date(Ticket.created_at))
		.order_by("date")
		.all()
	)
	
	# Horas por dia (apenas de tickets válidos)
	daily_hours = (
		TimeEntry.query.with_entities(
			func.date(TimeEntry.created_at).label("date"),
			func.sum(TimeEntry.hours).label("total_hours")
		)
		.join(Ticket, TimeEntry.ticket_id == Ticket.id)
		.filter(
			TimeEntry.created_at >= start_dt, 
			TimeEntry.created_at <= end_dt,
			Ticket.status != "cancelado"
		)
		.group_by(func.date(TimeEntry.created_at))
		.order_by("date")
		.all()
	)
	
	# Tempo médio de resolução (simplificado)
	avg_resolution_time = 0  # Removido temporariamente devido a problemas de cache
	
	# Taxa de fechamento
	total_tickets = Ticket.query.filter(
		Ticket.created_at >= start_dt,
		Ticket.created_at <= end_dt
	).count()
	
	closed_tickets = Ticket.query.filter(
		Ticket.status == "fechado",
		Ticket.created_at >= start_dt,
		Ticket.created_at <= end_dt
	).count()
	
	closure_rate = (closed_tickets / total_tickets * 100) if total_tickets > 0 else 0
	
	return jsonify({
		"daily_tickets": [{"date": str(date), "count": count} for date, count in daily_tickets],
		"daily_hours": [{"date": str(date), "hours": float(hours or 0)} for date, hours in daily_hours],
		"avg_resolution_time": float(avg_resolution_time or 0),
		"closure_rate": float(closure_rate),
		"total_tickets": total_tickets,
		"closed_tickets": closed_tickets
	})


@bp.route("/api/service-performance")
@login_required
def api_service_performance():
	"""API para performance por tipo de serviço"""
	start = request.args.get("start")
	end = request.args.get("end")
	
	query = Ticket.query.join(Service, Ticket.service_id == Service.id)
	
	if start:
		start_dt = datetime.fromisoformat(start)
		query = query.filter(Ticket.created_at >= start_dt)
	if end:
		# Corrigir: adicionar 23:59:59 para incluir todo o dia
		end_dt = datetime.fromisoformat(end) + timedelta(hours=23, minutes=59, seconds=59)
		query = query.filter(Ticket.created_at <= end_dt)
	
	# Separar contagem de tickets e receita para evitar duplicação
	tickets_results = (
		query.with_entities(
			Service.id,
			Service.name,
			Service.hourly_rate,
			func.count(Ticket.id).label("tickets_count"),
			func.sum(Ticket.total_cost).filter(Ticket.status != "cancelado").label("total_revenue")
		)
		.group_by(Service.id, Service.name, Service.hourly_rate)
		.all()
	)
	
	# Calcular horas separadamente
	hours_results = (
		query.with_entities(
			Service.id,
			func.sum(TimeEntry.hours).label("total_hours"),
			func.avg(TimeEntry.hours).label("avg_hours_per_ticket")
		)
		.outerjoin(TimeEntry, Ticket.id == TimeEntry.ticket_id)
		.filter(Ticket.status != "cancelado")
		.group_by(Service.id)
		.all()
	)
	
	# Combinar resultados
	hours_dict = {service_id: (total_hours, avg_hours) for service_id, total_hours, avg_hours in hours_results}
	results = []
	for service_id, name, hourly_rate, tickets_count, total_revenue in tickets_results:
		total_hours, avg_hours = hours_dict.get(service_id, (0, 0))
		results.append((service_id, name, hourly_rate, tickets_count, total_hours, avg_hours, total_revenue))
	
	data = []
	for service_id, name, hourly_rate, tickets_count, total_hours, avg_hours, total_revenue in results:
		data.append({
			"service_id": service_id,
			"name": name,
			"hourly_rate": float(hourly_rate or 0),
			"tickets_count": tickets_count or 0,
			"total_hours": float(total_hours or 0),
			"avg_hours_per_ticket": float(avg_hours or 0),
			"total_revenue": float(total_revenue or 0)
		})
	
	return jsonify(data)


# Rotas legadas para compatibilidade
@bp.route("/tickets")
@login_required
def tickets_report():
	"""Relatório legado de tickets"""
	start = request.args.get("start")
	end = request.args.get("end")
	user_id = request.args.get("user_id", type=int)
	client_id = request.args.get("client_id", type=int)

	query = Ticket.query
	if user_id:
		query = query.filter(Ticket.assigned_to_id == user_id)
	if client_id:
		query = query.filter(Ticket.client_id == client_id)
	if start:
		start_dt = datetime.fromisoformat(start)
		query = query.filter(Ticket.created_at >= start_dt)
	if end:
		# Corrigir: adicionar 23:59:59 para incluir todo o dia
		end_dt = datetime.fromisoformat(end) + timedelta(hours=23, minutes=59, seconds=59)
		query = query.filter(Ticket.created_at <= end_dt)

	tickets = query.order_by(Ticket.created_at.desc()).all()
	users = User.query.filter_by(status="1").order_by(User.name.asc()).all()
	clients = Client.query.all()
	return render_template("reports/tickets.html", tickets=tickets, users=users, clients=clients)


@bp.route("/horas")
@login_required
def hours_report():
	"""Relatório legado de horas"""
	start = request.args.get("start")
	end = request.args.get("end")
	user_id = request.args.get("user_id", type=int)
	client_id = request.args.get("client_id", type=int)

	query = TimeEntry.query.join(TimeEntry.ticket)
	if user_id:
		query = query.filter(TimeEntry.user_id == user_id)
	if client_id:
		query = query.filter(Ticket.client_id == client_id)
	if start:
		start_dt = datetime.fromisoformat(start)
		query = query.filter(TimeEntry.created_at >= start_dt)
	if end:
		# Corrigir: adicionar 23:59:59 para incluir todo o dia
		end_dt = datetime.fromisoformat(end) + timedelta(hours=23, minutes=59, seconds=59)
		query = query.filter(TimeEntry.created_at <= end_dt)

	summary = (
		query.with_entities(
			func.coalesce(TimeEntry.user_id, 0).label("user_id"),
			func.sum(TimeEntry.hours).label("total_hours"),
		)
		.group_by(TimeEntry.user_id)
		.all()
	)
	users = {u.id: u for u in User.query.filter_by(status="1").all()}
	clients = Client.query.all()
	return render_template("reports/hours.html", summary=summary, users=users, clients=clients)


@bp.route("/export/billing-by-technician")
@login_required
def export_billing_by_technician():
	"""Exportar relatório de faturamento por técnico para Excel"""
	start = request.args.get("start")
	end = request.args.get("end")
	
	# Buscar dados de faturamento (reutilizar lógica da API)
	# 1. FATURAMENTO DE TICKETS
	tickets_query = Ticket.query.join(User, Ticket.assigned_to_id == User.id)
	tickets_query = tickets_query.filter(Ticket.status != "cancelado")
	tickets_query = tickets_query.filter(Ticket.status == "fechado")
	tickets_query = tickets_query.filter(Ticket.closed_at.isnot(None))
	tickets_query = tickets_query.filter(User.status == "1")
	
	if start:
		start_dt = datetime.fromisoformat(start)
		tickets_query = tickets_query.filter(Ticket.closed_at >= start_dt)
	if end:
		end_dt = datetime.fromisoformat(end) + timedelta(hours=23, minutes=59, seconds=59)
		tickets_query = tickets_query.filter(Ticket.closed_at <= end_dt)
	
	tickets_results = (
		tickets_query.with_entities(
			Ticket.assigned_to_id,
			User.name,
			User.role,
			func.sum(Ticket.total_cost).label("tickets_billing"),
			func.count(Ticket.id).label("tickets_count")
		)
		.group_by(Ticket.assigned_to_id, User.name, User.role)
		.all()
	)
	
	# 2. FATURAMENTO DE ORDENS DE SERVIÇO
	service_orders_query = ServiceOrder.query.join(User, ServiceOrder.technician_id == User.id)
	service_orders_query = service_orders_query.filter(ServiceOrder.value > 0)
	service_orders_query = service_orders_query.filter(User.status == "1")
	
	if start:
		start_dt = datetime.fromisoformat(start)
		service_orders_query = service_orders_query.filter(ServiceOrder.completion_date >= start_dt)
	if end:
		end_dt = datetime.fromisoformat(end) + timedelta(hours=23, minutes=59, seconds=59)
		service_orders_query = service_orders_query.filter(ServiceOrder.completion_date <= end_dt)
	
	service_orders_results = (
		service_orders_query.with_entities(
			ServiceOrder.technician_id,
			User.name,
			User.role,
			func.sum(ServiceOrder.value).label("service_orders_billing"),
			func.count(ServiceOrder.id).label("service_orders_count")
		)
		.group_by(ServiceOrder.technician_id, User.name, User.role)
		.all()
	)
	
	# 3. COMBINAR RESULTADOS
	tickets_dict = {}
	for user_id, name, role, tickets_billing, tickets_count in tickets_results:
		tickets_dict[user_id] = {
			"name": name,
			"role": role,
			"tickets_billing": float(tickets_billing or 0),
			"tickets_count": tickets_count or 0
		}
	
	service_orders_dict = {}
	for user_id, name, role, service_orders_billing, service_orders_count in service_orders_results:
		service_orders_dict[user_id] = {
			"name": name,
			"role": role,
			"service_orders_billing": float(service_orders_billing or 0),
			"service_orders_count": service_orders_count or 0
		}
	
	all_user_ids = set(tickets_dict.keys()) | set(service_orders_dict.keys())
	
	# Criar workbook Excel
	wb = Workbook()
	ws = wb.active
	ws.title = "Faturamento por Técnico"
	
	# Estilos
	header_font = Font(bold=True, color="FFFFFF")
	header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
	subheader_font = Font(bold=True, color="000000")
	subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
	border = Border(
		left=Side(style='thin'),
		right=Side(style='thin'),
		top=Side(style='thin'),
		bottom=Side(style='thin')
	)
	center_alignment = Alignment(horizontal='center', vertical='center')
	
	# Cabeçalho principal
	row = 1
	ws.merge_cells(f'A{row}:H{row}')
	ws[f'A{row}'] = f"RELATÓRIO DE FATURAMENTO POR TÉCNICO"
	ws[f'A{row}'].font = Font(bold=True, size=16)
	ws[f'A{row}'].alignment = center_alignment
	
	row += 1
	ws.merge_cells(f'A{row}:H{row}')
	period_text = f"Período: {start} a {end}" if start and end else "Período: Todos os dados"
	ws[f'A{row}'] = period_text
	ws[f'A{row}'].font = Font(bold=True, size=12)
	ws[f'A{row}'].alignment = center_alignment
	
	row += 2
	
	# Resumo por técnico
	ws[f'A{row}'] = "RESUMO POR TÉCNICO"
	ws[f'A{row}'].font = subheader_font
	ws[f'A{row}'].fill = subheader_fill
	ws.merge_cells(f'A{row}:H{row}')
	
	row += 1
	# Cabeçalhos do resumo
	headers = ["Técnico", "Função", "Tickets", "Faturamento Tickets", "Ordens", "Faturamento Ordens", "Total Itens", "Faturamento Total"]
	for col, header in enumerate(headers, 1):
		cell = ws.cell(row=row, column=col, value=header)
		cell.font = header_font
		cell.fill = header_fill
		cell.alignment = center_alignment
		cell.border = border
	
	row += 1
	
	# Dados do resumo
	for user_id in sorted(all_user_ids, key=lambda x: (tickets_dict.get(x, {}).get("tickets_billing", 0) + service_orders_dict.get(x, {}).get("service_orders_billing", 0)), reverse=True):
		tickets_data = tickets_dict.get(user_id, {"tickets_billing": 0, "tickets_count": 0})
		service_orders_data = service_orders_dict.get(user_id, {"service_orders_billing": 0, "service_orders_count": 0})
		user_data = tickets_data if user_id in tickets_dict else service_orders_data
		
		total_billing = tickets_data["tickets_billing"] + service_orders_data["service_orders_billing"]
		total_items = tickets_data["tickets_count"] + service_orders_data["service_orders_count"]
		
		ws.cell(row=row, column=1, value=user_data["name"]).border = border
		ws.cell(row=row, column=2, value=user_data["role"]).border = border
		ws.cell(row=row, column=3, value=tickets_data["tickets_count"]).border = border
		ws.cell(row=row, column=4, value=f"R$ {tickets_data['tickets_billing']:.2f}").border = border
		ws.cell(row=row, column=5, value=service_orders_data["service_orders_count"]).border = border
		ws.cell(row=row, column=6, value=f"R$ {service_orders_data['service_orders_billing']:.2f}").border = border
		ws.cell(row=row, column=7, value=total_items).border = border
		ws.cell(row=row, column=8, value=f"R$ {total_billing:.2f}").border = border
		
		row += 1
	
	row += 2
	
	# Detalhes por técnico
	for user_id in sorted(all_user_ids, key=lambda x: (tickets_dict.get(x, {}).get("tickets_billing", 0) + service_orders_dict.get(x, {}).get("service_orders_billing", 0)), reverse=True):
		user_data = tickets_dict.get(user_id, service_orders_dict.get(user_id, {}))
		
		# Cabeçalho do técnico
		ws[f'A{row}'] = f"DETALHES - {user_data['name']}"
		ws[f'A{row}'].font = subheader_font
		ws[f'A{row}'].fill = subheader_fill
		ws.merge_cells(f'A{row}:H{row}')
		
		row += 1
		
		# Buscar tickets detalhados
		tickets_detail_query = Ticket.query.filter(Ticket.assigned_to_id == user_id)
		tickets_detail_query = tickets_detail_query.filter(Ticket.status == "fechado")
		tickets_detail_query = tickets_detail_query.filter(Ticket.closed_at.isnot(None))
		tickets_detail_query = tickets_detail_query.filter(Ticket.total_cost > 0)
		
		if start:
			start_dt = datetime.fromisoformat(start)
			tickets_detail_query = tickets_detail_query.filter(Ticket.closed_at >= start_dt)
		if end:
			end_dt = datetime.fromisoformat(end) + timedelta(hours=23, minutes=59, seconds=59)
			tickets_detail_query = tickets_detail_query.filter(Ticket.closed_at <= end_dt)
		
		tickets_detail = tickets_detail_query.order_by(Ticket.closed_at.desc()).all()
		
		# Buscar ordens de serviço detalhadas
		service_orders_detail_query = ServiceOrder.query.filter(ServiceOrder.technician_id == user_id)
		service_orders_detail_query = service_orders_detail_query.filter(ServiceOrder.value > 0)
		
		if start:
			start_dt = datetime.fromisoformat(start)
			service_orders_detail_query = service_orders_detail_query.filter(ServiceOrder.completion_date >= start_dt)
		if end:
			end_dt = datetime.fromisoformat(end) + timedelta(hours=23, minutes=59, seconds=59)
			service_orders_detail_query = service_orders_detail_query.filter(ServiceOrder.completion_date <= end_dt)
		
		service_orders_detail = service_orders_detail_query.order_by(ServiceOrder.completion_date.desc()).all()
		
		# TICKETS
		if tickets_detail:
			ws[f'A{row}'] = "TICKETS FATURADOS"
			ws[f'A{row}'].font = Font(bold=True, italic=True)
			ws.merge_cells(f'A{row}:H{row}')
			row += 1
			
			# Cabeçalhos dos tickets
			ticket_headers = ["ID", "Título", "Cliente", "Serviço", "Horas", "Data Fechamento", "Valor", "Tipo"]
			for col, header in enumerate(ticket_headers, 1):
				cell = ws.cell(row=row, column=col, value=header)
				cell.font = header_font
				cell.fill = header_fill
				cell.alignment = center_alignment
				cell.border = border
			
			row += 1
			
			# Dados dos tickets
			for ticket in tickets_detail:
				ws.cell(row=row, column=1, value=f"#{ticket.id}").border = border
				ws.cell(row=row, column=2, value=ticket.title).border = border
				ws.cell(row=row, column=3, value=ticket.display_client_name()).border = border
				ws.cell(row=row, column=4, value=ticket.service.name if ticket.service else "N/A").border = border
				ws.cell(row=row, column=5, value=f"{ticket.total_hours():.1f}h").border = border
				ws.cell(row=row, column=6, value=ticket.closed_at.strftime("%d/%m/%Y") if ticket.closed_at else "").border = border
				ws.cell(row=row, column=7, value=f"R$ {ticket.total_cost:.2f}").border = border
				ws.cell(row=row, column=8, value="Ticket").border = border
				row += 1
		
		# ORDENS DE SERVIÇO
		if service_orders_detail:
			if tickets_detail:
				row += 1
			
			ws[f'A{row}'] = "ORDENS DE SERVIÇO"
			ws[f'A{row}'].font = Font(bold=True, italic=True)
			ws.merge_cells(f'A{row}:H{row}')
			row += 1
			
			# Cabeçalhos das ordens
			order_headers = ["Código", "Serviço Executado", "Cliente", "Data Finalização", "Valor", "Tipo", "", ""]
			for col, header in enumerate(order_headers, 1):
				cell = ws.cell(row=row, column=col, value=header)
				cell.font = header_font
				cell.fill = header_fill
				cell.alignment = center_alignment
				cell.border = border
			
			row += 1
			
			# Dados das ordens
			for order in service_orders_detail:
				ws.cell(row=row, column=1, value=order.codigo).border = border
				ws.cell(row=row, column=2, value=order.service_executed).border = border
				ws.cell(row=row, column=3, value=order.client_name).border = border
				ws.cell(row=row, column=4, value=order.completion_date.strftime("%d/%m/%Y") if order.completion_date else "").border = border
				ws.cell(row=row, column=5, value=f"R$ {order.value:.2f}").border = border
				ws.cell(row=row, column=6, value="Ordem de Serviço").border = border
				ws.cell(row=row, column=7, value="").border = border
				ws.cell(row=row, column=8, value="").border = border
				row += 1
		
		row += 2
	
	# Ajustar largura das colunas
	for col in range(1, 9):
		ws.column_dimensions[get_column_letter(col)].width = 15
	
	# Salvar em memória
	output = io.BytesIO()
	wb.save(output)
	output.seek(0)
	
	# Nome do arquivo
	filename = f"faturamento_por_tecnico_{start}_{end}.xlsx" if start and end else "faturamento_por_tecnico.xlsx"
	
	return send_file(
		output,
		as_attachment=True,
		download_name=filename,
		mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
	)


def generate_hours_by_client_pdf(data):
	"""Gera PDF com relatório detalhado de horas por cliente"""
	buffer = io.BytesIO()
	doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
	
	# Estilos
	styles = getSampleStyleSheet()
	
	# Estilo personalizado para título
	title_style = ParagraphStyle(
		'CustomTitle',
		parent=styles['Heading1'],
		fontSize=18,
		spaceAfter=30,
		alignment=TA_CENTER,
		textColor=colors.HexColor('#366092')
	)
	
	# Estilo para subtítulos
	subtitle_style = ParagraphStyle(
		'CustomSubtitle',
		parent=styles['Heading2'],
		fontSize=14,
		spaceAfter=12,
		spaceBefore=20,
		textColor=colors.HexColor('#2c5282')
	)
	
	# Estilo para informações do cliente
	client_info_style = ParagraphStyle(
		'ClientInfo',
		parent=styles['Normal'],
		fontSize=12,
		spaceAfter=6,
		textColor=colors.HexColor('#4a5568')
	)
	
	# Estilo para resumo
	summary_style = ParagraphStyle(
		'Summary',
		parent=styles['Normal'],
		fontSize=11,
		spaceAfter=8,
		textColor=colors.HexColor('#2d3748')
	)
	
	# Estilo para cabeçalhos de tabela
	header_style = ParagraphStyle(
		'TableHeader',
		parent=styles['Normal'],
		fontSize=10,
		fontName='Helvetica-Bold',
		textColor=colors.white,
		alignment=TA_CENTER
	)
	
	# Estilo para dados da tabela
	data_style = ParagraphStyle(
		'TableData',
		parent=styles['Normal'],
		fontSize=9,
		textColor=colors.HexColor('#2d3748')
	)
	
	# Conteúdo do PDF
	story = []
	
	# Título principal
	story.append(Paragraph("RELATÓRIO DE HORAS POR CLIENTE", title_style))
	story.append(Spacer(1, 20))
	
	# Informações do cliente
	client_info = data['client_info']
	story.append(Paragraph("INFORMAÇÕES DO CLIENTE", subtitle_style))
	story.append(Paragraph(f"<b>Nome:</b> {client_info['name']}", client_info_style))
	story.append(Paragraph(f"<b>Tipo:</b> {client_info['type']}", client_info_style))
	if client_info.get('document'):
		story.append(Paragraph(f"<b>Documento:</b> {client_info['document']}", client_info_style))
	if client_info.get('phone'):
		story.append(Paragraph(f"<b>Telefone:</b> {client_info['phone']}", client_info_style))
	
	story.append(Spacer(1, 20))
	
	# Período
	period = data['period']
	story.append(Paragraph("PERÍODO ANALISADO", subtitle_style))
	if period['start'] and period['end']:
		start_date = datetime.fromisoformat(period['start']).strftime('%d/%m/%Y')
		end_date = datetime.fromisoformat(period['end']).strftime('%d/%m/%Y')
		story.append(Paragraph(f"<b>De:</b> {start_date} <b>Até:</b> {end_date}", client_info_style))
	elif period['start']:
		start_date = datetime.fromisoformat(period['start']).strftime('%d/%m/%Y')
		story.append(Paragraph(f"<b>A partir de:</b> {start_date}", client_info_style))
	elif period['end']:
		end_date = datetime.fromisoformat(period['end']).strftime('%d/%m/%Y')
		story.append(Paragraph(f"<b>Até:</b> {end_date}", client_info_style))
	else:
		story.append(Paragraph("<b>Todos os períodos</b>", client_info_style))
	
	story.append(Spacer(1, 20))
	
	# Resumo executivo
	totals = data['totals']
	story.append(Paragraph("RESUMO EXECUTIVO", subtitle_style))
	story.append(Paragraph(f"<b>Total de Horas:</b> {totals['total_hours']:.1f}h", summary_style))
	story.append(Paragraph(f"<b>Total de Tickets:</b> {totals['total_tickets']}", summary_style))
	story.append(Paragraph(f"<b>Total de Apontamentos:</b> {totals['total_entries']}", summary_style))
	if totals['total_tickets'] > 0:
		avg_hours_per_ticket = totals['total_hours'] / totals['total_tickets']
		story.append(Paragraph(f"<b>Média de Horas por Ticket:</b> {avg_hours_per_ticket:.1f}h", summary_style))
	
	story.append(Spacer(1, 30))
	
	# Detalhamento por ticket
	story.append(Paragraph("DETALHAMENTO POR TICKET", subtitle_style))
	
	tickets = data['tickets']
	if tickets:
		for i, ticket in enumerate(tickets):
			# Cabeçalho do ticket
			story.append(Paragraph(f"<b>Ticket #{ticket['id']} - {ticket['title']}</b>", data_style))
			
			# Informações básicas do ticket
			ticket_info = [
				f"<b>Status:</b> {ticket['status'].title()}",
				f"<b>Serviço:</b> {ticket['service_name']}",
				f"<b>Técnico:</b> {ticket['assigned_to']}",
				f"<b>Total de Horas:</b> {ticket['total_hours']:.1f}h"
			]
			
			if ticket['created_at']:
				created_date = datetime.fromisoformat(ticket['created_at']).strftime('%d/%m/%Y %H:%M')
				ticket_info.append(f"<b>Criado em:</b> {created_date}")
			
			if ticket['closed_at']:
				closed_date = datetime.fromisoformat(ticket['closed_at']).strftime('%d/%m/%Y %H:%M')
				ticket_info.append(f"<b>Fechado em:</b> {closed_date}")
			
			story.append(Paragraph(" | ".join(ticket_info), data_style))
			
			# Descrição do ticket
			if ticket['description']:
				# Exibir descrição completa (sem truncar) para não cortar informações
				description_text = ticket['description'].replace('\n', '<br/>')
				story.append(Paragraph(f"<b>Descrição:</b> {description_text}", data_style))
			
			story.append(Spacer(1, 10))
			
			# Pontamentos do ticket
			time_entries = ticket['time_entries']
			if time_entries:
				story.append(Paragraph("<b>Apontamentos:</b>", data_style))
				
				# Tabela de pontamentos
				table_data = [['Data/Hora', 'Técnico', 'Descrição', 'Horas', 'Cobrança']]
				
				# Definir estilos para colunas da tabela
				cell_style = ParagraphStyle(
					'CellStyle',
					parent=styles['Normal'],
					fontSize=8,
					leading=10,
					alignment=TA_CENTER
				)
				
				left_cell_style = ParagraphStyle(
					'LeftCellStyle',
					parent=styles['Normal'],
					fontSize=8,
					leading=10,
					alignment=TA_LEFT
				)

				for entry in time_entries:
					entry_date = datetime.fromisoformat(entry['created_at']).strftime('%d/%m/%Y\n%H:%M') if entry['created_at'] else 'N/A'
					charge_status = 'Não cobrar' if entry['no_charge'] else 'Cobrar'
					
					# Usar Paragraph para permitir quebra de linha automática
					description_p = Paragraph(entry['description'] or "", left_cell_style)
					technician_p = Paragraph(entry['user_name'], cell_style)
					
					table_data.append([
						entry_date,
						technician_p,
						description_p,
						f"{entry['hours']:.1f}h",
						charge_status
					])
				
				# Criar tabela com larguras ajustadas (Total: 17cm - margem segura)
				# Antes: [3*cm, 3*cm, 6*cm, 2*cm, 2*cm] = 16cm
				# Agora: [2.5*cm, 3*cm, 8.5*cm, 1.5*cm, 1.5*cm] = 17cm
				table = Table(table_data, colWidths=[2.5*cm, 3*cm, 8.5*cm, 1.5*cm, 1.5*cm])
				table.setStyle(TableStyle([
					('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
					('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
					('ALIGN', (0, 0), (-1, -1), 'CENTER'),
					('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
					('FONTSIZE', (0, 0), (-1, 0), 9),
					('BOTTOMPADDING', (0, 0), (-1, 0), 12),
					('BACKGROUND', (0, 1), (-1, -1), colors.beige),
					('GRID', (0, 0), (-1, -1), 1, colors.black),
					('FONTSIZE', (0, 1), (-1, -1), 8),
					('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
					# Alinhar descrição à esquerda
					('ALIGN', (2, 1), (2, -1), 'LEFT'),
				]))
				
				story.append(table)
			else:
				story.append(Paragraph("<i>Nenhum pontamento registrado</i>", data_style))
			
			story.append(Spacer(1, 20))
			
			# Quebra de página a cada 3 tickets para evitar páginas muito longas
			if (i + 1) % 3 == 0 and i < len(tickets) - 1:
				story.append(PageBreak())
	else:
		story.append(Paragraph("<i>Nenhum ticket encontrado para o período selecionado</i>", data_style))
	
	# Rodapé com data de geração
	story.append(Spacer(1, 30))
	story.append(Paragraph(f"<i>Relatório gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}</i>", data_style))
	
	# Construir PDF
	doc.build(story)
	buffer.seek(0)
	return buffer


def generate_hours_by_client_synthetic_pdf(data):
	"""Gera PDF sintético com relatório resumido de horas por cliente"""
	buffer = io.BytesIO()
	doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
	
	# Estilos
	styles = getSampleStyleSheet()
	
	# Estilo personalizado para título
	title_style = ParagraphStyle(
		'CustomTitle',
		parent=styles['Heading1'],
		fontSize=18,
		spaceAfter=30,
		alignment=TA_CENTER,
		textColor=colors.HexColor('#366092')
	)
	
	# Estilo para subtítulos
	subtitle_style = ParagraphStyle(
		'CustomSubtitle',
		parent=styles['Heading2'],
		fontSize=14,
		spaceAfter=12,
		spaceBefore=20,
		textColor=colors.HexColor('#2c5282')
	)
	
	# Estilo para informações do cliente
	client_info_style = ParagraphStyle(
		'ClientInfo',
		parent=styles['Normal'],
		fontSize=12,
		spaceAfter=6,
		textColor=colors.HexColor('#4a5568')
	)
	
	# Estilo para resumo
	summary_style = ParagraphStyle(
		'Summary',
		parent=styles['Normal'],
		fontSize=11,
		spaceAfter=8,
		textColor=colors.HexColor('#2d3748')
	)
	
	# Estilo para dados da tabela
	data_style = ParagraphStyle(
		'TableData',
		parent=styles['Normal'],
		fontSize=9,
		textColor=colors.HexColor('#2d3748')
	)
	
	# Conteúdo do PDF
	story = []
	
	# Título principal
	story.append(Paragraph("RELATÓRIO SINTÉTICO DE HORAS POR CLIENTE", title_style))
	story.append(Spacer(1, 20))
	
	# Informações do cliente
	client_info = data['client_info']
	story.append(Paragraph("INFORMAÇÕES DO CLIENTE", subtitle_style))
	story.append(Paragraph(f"<b>Nome:</b> {client_info['name']}", client_info_style))
	story.append(Paragraph(f"<b>Tipo:</b> {client_info['type']}", client_info_style))
	if client_info.get('document'):
		story.append(Paragraph(f"<b>Documento:</b> {client_info['document']}", client_info_style))
	if client_info.get('phone'):
		story.append(Paragraph(f"<b>Telefone:</b> {client_info['phone']}", client_info_style))
	
	story.append(Spacer(1, 20))
	
	# Período
	period = data['period']
	story.append(Paragraph("PERÍODO ANALISADO", subtitle_style))
	if period['start'] and period['end']:
		start_date = datetime.fromisoformat(period['start']).strftime('%d/%m/%Y')
		end_date = datetime.fromisoformat(period['end']).strftime('%d/%m/%Y')
		story.append(Paragraph(f"<b>De:</b> {start_date} <b>Até:</b> {end_date}", client_info_style))
	elif period['start']:
		start_date = datetime.fromisoformat(period['start']).strftime('%d/%m/%Y')
		story.append(Paragraph(f"<b>A partir de:</b> {start_date}", client_info_style))
	elif period['end']:
		end_date = datetime.fromisoformat(period['end']).strftime('%d/%m/%Y')
		story.append(Paragraph(f"<b>Até:</b> {end_date}", client_info_style))
	else:
		story.append(Paragraph("<b>Todos os períodos</b>", client_info_style))
	
	story.append(Spacer(1, 20))
	
	# Resumo executivo
	totals = data['totals']
	story.append(Paragraph("RESUMO EXECUTIVO", subtitle_style))
	story.append(Paragraph(f"<b>Total de Horas:</b> {totals['total_hours']:.1f}h", summary_style))
	story.append(Paragraph(f"<b>Total de Tickets:</b> {totals['total_tickets']}", summary_style))
	if totals['total_tickets'] > 0:
		avg_hours_per_ticket = totals['total_hours'] / totals['total_tickets']
		story.append(Paragraph(f"<b>Média de Horas por Ticket:</b> {avg_hours_per_ticket:.1f}h", summary_style))
	
	story.append(Spacer(1, 30))
	
	# Tabela sintética de tickets
	story.append(Paragraph("RESUMO POR TICKET", subtitle_style))
	
	tickets = data['tickets']
	if tickets:
		# Cabeçalho da tabela
		table_data = [['Ticket', 'Título', 'Status', 'Serviço', 'Técnico', 'Criação', 'Fechamento', 'Hrs']]
		
		# Estilos para células
		cell_style = ParagraphStyle(
			'CellStyle',
			parent=styles['Normal'],
			fontSize=7,
			leading=8,
			alignment=TA_CENTER
		)
		
		left_cell_style = ParagraphStyle(
			'LeftCellStyle',
			parent=styles['Normal'],
			fontSize=7,
			leading=8,
			alignment=TA_LEFT
		)
		
		for ticket in tickets:
			created_date = datetime.fromisoformat(ticket['created_at']).strftime('%d/%m/%y') if ticket['created_at'] else 'N/A'
			closed_date = datetime.fromisoformat(ticket['closed_at']).strftime('%d/%m/%y') if ticket['closed_at'] else '-'
			
			# Usar Paragraph para wrapping
			title_p = Paragraph(ticket['title'] or "", left_cell_style)
			service_p = Paragraph(ticket['service_name'] or "", left_cell_style)
			technician_p = Paragraph(ticket['assigned_to'] or "", left_cell_style)
			
			table_data.append([
				f"#{ticket['id']}",
				title_p,
				ticket['status'].title(),
				service_p,
				technician_p,
				created_date,
				closed_date,
				f"{ticket['total_hours']:.1f}"
			])
		
		# Criar tabela com larguras ajustadas (Total ~16.7cm para caber em 17cm)
		# ID, Título, Status, Serviço, Técnico, Criação, Fechamento, Horas
		col_widths = [1.3*cm, 4.0*cm, 1.7*cm, 2.7*cm, 2.2*cm, 1.9*cm, 1.9*cm, 1.0*cm]
		
		table = Table(table_data, colWidths=col_widths)
		table.setStyle(TableStyle([
			('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
			('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
			('ALIGN', (0, 0), (-1, -1), 'CENTER'),
			('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
			('FONTSIZE', (0, 0), (-1, 0), 8),
			('BOTTOMPADDING', (0, 0), (-1, 0), 12),
			('BACKGROUND', (0, 1), (-1, -1), colors.beige),
			('GRID', (0, 0), (-1, -1), 1, colors.black),
			('FONTSIZE', (0, 1), (-1, -1), 7),
			('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
			('ALIGN', (1, 1), (1, -1), 'LEFT'),  # Alinhar título à esquerda
			('ALIGN', (3, 1), (3, -1), 'LEFT'),  # Alinhar serviço à esquerda
			('ALIGN', (4, 1), (4, -1), 'LEFT'),  # Alinhar técnico à esquerda
		]))
		
		story.append(table)
		
		# Resumo por status
		story.append(Spacer(1, 20))
		story.append(Paragraph("RESUMO POR STATUS", subtitle_style))
		
		status_summary = {}
		for ticket in tickets:
			status = ticket['status']
			if status not in status_summary:
				status_summary[status] = {'count': 0, 'hours': 0}
			status_summary[status]['count'] += 1
			status_summary[status]['hours'] += ticket['total_hours']
		
		status_table_data = [['Status', 'Quantidade', 'Total de Horas', 'Média por Ticket']]
		for status, data in status_summary.items():
			avg_hours = data['hours'] / data['count'] if data['count'] > 0 else 0
			status_table_data.append([
				status.title(),
				str(data['count']),
				f"{data['hours']:.1f}h",
				f"{avg_hours:.1f}h"
			])
		
		status_table = Table(status_table_data, colWidths=[3*cm, 2*cm, 3*cm, 3*cm])
		status_table.setStyle(TableStyle([
			('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c5282')),
			('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
			('ALIGN', (0, 0), (-1, -1), 'CENTER'),
			('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
			('FONTSIZE', (0, 0), (-1, 0), 9),
			('BOTTOMPADDING', (0, 0), (-1, 0), 12),
			('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
			('GRID', (0, 0), (-1, -1), 1, colors.black),
			('FONTSIZE', (0, 1), (-1, -1), 8),
			('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
		]))
		
		story.append(status_table)
		
	else:
		story.append(Paragraph("<i>Nenhum ticket encontrado para o período selecionado</i>", data_style))
	
	# Rodapé com data de geração
	story.append(Spacer(1, 30))
	story.append(Paragraph(f"<i>Relatório sintético gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}</i>", data_style))
	
	# Construir PDF
	doc.build(story)
	buffer.seek(0)
	return buffer


@bp.route("/export/hours-by-client-pdf")
@login_required
def export_hours_by_client_pdf():
	"""Exportar relatório detalhado de horas por cliente para PDF"""
	start = request.args.get("start")
	end = request.args.get("end")
	client_id = request.args.get("client_id", type=int)
	external_client_id = request.args.get("external_client_id", type=int)
	external_client_name = request.args.get("external_client_name")
	
	if not client_id and not external_client_id and not external_client_name:
		return jsonify({"error": "Cliente não especificado"}), 400
	
	# Buscar dados usando a API detalhada
	from flask import current_app
	with current_app.test_request_context():
		# Simular requisição para a API detalhada
		from flask import request as flask_request
		original_args = flask_request.args
		
		# Criar nova requisição com os parâmetros
		import urllib.parse
		params = {}
		if start:
			params['start'] = start
		if end:
			params['end'] = end
		if client_id:
			params['client_id'] = str(client_id)
		if external_client_id:
			params['external_client_id'] = str(external_client_id)
		if external_client_name:
			params['external_client_name'] = external_client_name
		
		# Fazer a requisição interna
		with current_app.test_client() as client:
			url = f"/relatorios/api/hours-by-client-detailed?{urllib.parse.urlencode(params)}"
			response = client.get(url)
			
			if response.status_code != 200:
				return jsonify({"error": "Erro ao buscar dados"}), 500
			
			data = response.get_json()
	
	# Gerar PDF
	try:
		pdf_buffer = generate_hours_by_client_pdf(data)
		
		# Nome do arquivo
		client_name = data['client_info']['name'].replace(' ', '_').replace('/', '_')
		start_str = start.replace('-', '') if start else 'inicio'
		end_str = end.replace('-', '') if end else 'fim'
		filename = f"relatorio_horas_{client_name}_{start_str}_{end_str}.pdf"
		
		return send_file(
			pdf_buffer,
			as_attachment=True,
			download_name=filename,
			mimetype='application/pdf'
		)
		
	except Exception as e:
		return jsonify({"error": f"Erro ao gerar PDF: {str(e)}"}), 500


@bp.route("/export/hours-by-client-synthetic-pdf")
@login_required
def export_hours_by_client_synthetic_pdf():
	"""Exportar relatório sintético de horas por cliente para PDF"""
	start = request.args.get("start")
	end = request.args.get("end")
	client_id = request.args.get("client_id", type=int)
	external_client_id = request.args.get("external_client_id", type=int)
	external_client_name = request.args.get("external_client_name")
	
	if not client_id and not external_client_id and not external_client_name:
		return jsonify({"error": "Cliente não especificado"}), 400
	
	# Buscar dados usando a API sintética
	from flask import current_app
	with current_app.test_request_context():
		# Criar nova requisição com os parâmetros
		import urllib.parse
		params = {}
		if start:
			params['start'] = start
		if end:
			params['end'] = end
		if client_id:
			params['client_id'] = str(client_id)
		if external_client_id:
			params['external_client_id'] = str(external_client_id)
		if external_client_name:
			params['external_client_name'] = external_client_name
		
		# Fazer a requisição interna
		with current_app.test_client() as client:
			url = f"/relatorios/api/hours-by-client-synthetic?{urllib.parse.urlencode(params)}"
			response = client.get(url)
			
			if response.status_code != 200:
				return jsonify({"error": "Erro ao buscar dados"}), 500
			
			data = response.get_json()
	
	# Gerar PDF sintético
	try:
		pdf_buffer = generate_hours_by_client_synthetic_pdf(data)
		
		# Nome do arquivo
		client_name = data['client_info']['name'].replace(' ', '_').replace('/', '_')
		start_str = start.replace('-', '') if start else 'inicio'
		end_str = end.replace('-', '') if end else 'fim'
		filename = f"relatorio_sintetico_horas_{client_name}_{start_str}_{end_str}.pdf"
		
		return send_file(
			pdf_buffer,
			as_attachment=True,
			download_name=filename,
			mimetype='application/pdf'
		)
		
	except Exception as e:
		return jsonify({"error": f"Erro ao gerar PDF sintético: {str(e)}"}), 500


@bp.route("/export/hours-by-client-excel")
@login_required
def export_hours_by_client_excel():
	"""Exportar relatório resumido de horas por cliente para Excel"""
	start = request.args.get("start")
	end = request.args.get("end")
	data = api_hours_by_client().get_json() or []
	rows = [
		[
			item.get("client_name"),
			item.get("client_type"),
			f"{float(item.get('total_hours') or 0):.1f}h",
			item.get("tickets_count") or 0,
			f"{float(item.get('avg_hours_per_entry') or 0):.1f}h",
		]
		for item in data
	]
	filename = f"horas_por_cliente_{start}_{end}.xlsx" if start and end else "horas_por_cliente.xlsx"
	return _xlsx_table(
		"RELATÓRIO DE HORAS POR CLIENTE",
		filename,
		["Cliente", "Tipo", "Total de Horas", "Tickets", "Média por Entrada"],
		rows,
		start,
		end,
	)


def _xlsx_table(title, filename, headers, rows, start=None, end=None):
	wb = Workbook()
	ws = wb.active
	ws.title = title[:31]
	header_font = Font(bold=True, color="FFFFFF")
	header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
	border = Border(
		left=Side(style="thin"),
		right=Side(style="thin"),
		top=Side(style="thin"),
		bottom=Side(style="thin"),
	)
	center = Alignment(horizontal="center", vertical="center")
	last = get_column_letter(len(headers))
	ws.merge_cells(f"A1:{last}1")
	ws["A1"] = title
	ws["A1"].font = Font(bold=True, size=16)
	ws["A1"].alignment = center
	ws.merge_cells(f"A2:{last}2")
	ws["A2"] = f"Período: {start} a {end}" if start and end else "Período: Todos os dados"
	ws["A2"].font = Font(bold=True, size=12)
	ws["A2"].alignment = center
	row = 4
	for col, header in enumerate(headers, 1):
		cell = ws.cell(row=row, column=col, value=header)
		cell.font = header_font
		cell.fill = header_fill
		cell.alignment = center
		cell.border = border
	for values in rows:
		row += 1
		for col, value in enumerate(values, 1):
			cell = ws.cell(row=row, column=col, value=value)
			cell.border = border
	for col in range(1, len(headers) + 1):
		ws.column_dimensions[get_column_letter(col)].width = 22
	output = io.BytesIO()
	wb.save(output)
	output.seek(0)
	return send_file(
		output,
		as_attachment=True,
		download_name=filename,
		mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
	)


def _period():
	return request.args.get("start"), request.args.get("end")


def _role_label(role):
	r = (role or "").lower()
	if r in ("admin", "administrador", "administrator"):
		return "Admin"
	if r == "tecnico":
		return "Técnico"
	if r == "viewer":
		return "Visualizador"
	return role or "—"


@bp.route("/export/hours-by-technician")
@login_required
def export_hours_by_technician():
	start, end = _period()
	data = api_hours_by_technician().get_json() or []
	rows = [
		[d.get("name"), _role_label(d.get("role")), round(float(d.get("total_hours") or 0), 1),
		 d.get("entries_count") or 0, d.get("tickets_count") or 0, round(float(d.get("avg_hours_per_entry") or 0), 1)]
		for d in data
	]
	name = f"horas_por_tecnico_{start}_{end}.xlsx" if start and end else "horas_por_tecnico.xlsx"
	return _xlsx_table(
		"RELATÓRIO DE HORAS POR TÉCNICO",
		name,
		["Técnico", "Função", "Total de Horas", "Entradas", "Tickets", "Média por Entrada"],
		rows,
		start,
		end,
	)


@bp.route("/export/tickets-by-technician")
@login_required
def export_tickets_by_technician():
	start, end = _period()
	data = api_tickets_by_technician().get_json() or []
	rows = [
		[d.get("name"), _role_label(d.get("role")), d.get("total_tickets") or 0, d.get("open_tickets") or 0,
		 d.get("in_progress_tickets") or 0, d.get("closed_tickets") or 0, round(float(d.get("total_hours") or 0), 1)]
		for d in data
	]
	name = f"tickets_por_tecnico_{start}_{end}.xlsx" if start and end else "tickets_por_tecnico.xlsx"
	return _xlsx_table(
		"RELATÓRIO DE TICKETS POR TÉCNICO",
		name,
		["Técnico", "Função", "Total", "Abertos", "Em atendimento", "Encerrados", "Horas"],
		rows,
		start,
		end,
	)


@bp.route("/export/tickets-by-client")
@login_required
def export_tickets_by_client():
	start, end = _period()
	data = api_tickets_by_client().get_json() or []
	rows = [
		[d.get("client_name"), d.get("client_type"), d.get("total_tickets") or 0, d.get("open_tickets") or 0,
		 d.get("in_progress_tickets") or 0, d.get("closed_tickets") or 0, round(float(d.get("total_hours") or 0), 1)]
		for d in data
	]
	name = f"tickets_por_cliente_{start}_{end}.xlsx" if start and end else "tickets_por_cliente.xlsx"
	return _xlsx_table(
		"RELATÓRIO DE TICKETS POR CLIENTE",
		name,
		["Cliente", "Tipo", "Total", "Abertos", "Em atendimento", "Encerrados", "Horas"],
		rows,
		start,
		end,
	)


@bp.route("/export/service-performance")
@login_required
def export_service_performance():
	start, end = _period()
	data = api_service_performance().get_json() or []
	rows = [
		[d.get("name"), float(d.get("hourly_rate") or 0), d.get("tickets_count") or 0,
		 round(float(d.get("total_hours") or 0), 1), round(float(d.get("avg_hours_per_ticket") or 0), 1),
		 float(d.get("total_revenue") or 0)]
		for d in data
	]
	name = f"performance_servicos_{start}_{end}.xlsx" if start and end else "performance_servicos.xlsx"
	return _xlsx_table(
		"RELATÓRIO DE PERFORMANCE POR SERVIÇO",
		name,
		["Serviço", "Taxa/hora", "Tickets", "Horas", "Média por ticket", "Receita"],
		rows,
		start,
		end,
	)


@bp.route("/export/productivity")
@login_required
def export_productivity():
	start, end = _period()
	data = api_productivity_metrics().get_json() or {}
	rows = [
		["Tickets no período", data.get("total_tickets") or 0],
		["Tickets fechados", data.get("closed_tickets") or 0],
		["Taxa de fechamento (%)", round(float(data.get("closure_rate") or 0), 1)],
	]
	for item in data.get("daily_tickets") or []:
		rows.append([f"Tickets em {item.get('date')}", item.get("count") or 0])
	for item in data.get("daily_hours") or []:
		rows.append([f"Horas em {item.get('date')}", round(float(item.get("hours") or 0), 1)])
	name = f"produtividade_{start}_{end}.xlsx" if start and end else "produtividade.xlsx"
	return _xlsx_table("RELATÓRIO DE PRODUTIVIDADE", name, ["Métrica", "Valor"], rows, start, end)
